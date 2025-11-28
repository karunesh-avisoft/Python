from faker import Faker
from openpyxl import Workbook

wb=Workbook()
ws=wb.active
fake = Faker()

# Create headers 
ws.append(["Name", "Email", "Address", "Phone"])

for i in range(1,11):
    for j in range(1,5):
        ws.cell(row=i,column=1).value=fake.name()
        ws.cell(row=i,column=2).value=fake.email()
        ws.cell(row=i,column=3).value=fake.address()
        ws.cell(row=i,column=4).value=fake.phone_number()
wb.save('test_data.xlsx')